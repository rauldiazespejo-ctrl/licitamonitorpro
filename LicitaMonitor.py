"""
LICITAMONITOR — Soldesp
========================
Monitor multi-portal de licitaciones industriales.
Portales soportados: Wherex · Portal Minero

REQUISITOS:
    pip install selenium openpyxl webdriver-manager requests

CONFIGURACIÓN:
    Editar config.json con credenciales de cada portal.
"""

import json
import logging
import os
import re
import smtplib
import sys
import time
from abc import ABC, abstractmethod
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path

from selenium import webdriver
from selenium.common.exceptions import (
    NoSuchElementException,
    StaleElementReferenceException,
    TimeoutException,
    WebDriverException,
)
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════
# PALETA OSCURA CORPORATIVA
# ═══════════════════════════════════════════════════════════════

C_HEADER_BG      = "0A1628"   # Casi negro azulado — cabeceras principales
C_SUBHEADER_BG   = "0F2236"   # Azul muy oscuro — subtítulos
C_ROW_EVEN       = "112033"   # Filas pares
C_ROW_ODD        = "0D1A2B"   # Filas impares
C_RESUMEN_BG     = "0F1E30"   # Fondo hoja resumen
C_RESUMEN_ROW    = "162840"   # Filas hoja resumen

C_BADGE_WHEREX   = "1255A3"   # Azul — badge portal Wherex
C_BADGE_MINERO   = "6B1515"   # Rojo oscuro — badge Portal Minero

C_TEXT_HEADER    = "FFFFFF"   # Blanco puro — texto en cabeceras
C_TEXT_DATA      = "C8DCF0"   # Azul muy claro — texto en celdas
C_TEXT_MUTED     = "5A7A99"   # Gris azulado — texto secundario
C_LINK           = "4FC3F7"   # Azul eléctrico — hipervínculos
C_BORDER         = "1A3550"   # Borde oscuro


# ═══════════════════════════════════════════════════════════════
# CONFIGURACIÓN
# ═══════════════════════════════════════════════════════════════

def cargar_config() -> dict:
    """Carga config.json desde el directorio del ejecutable o script."""
    base = Path(sys.executable).parent if getattr(sys, "frozen", False) else Path(__file__).parent
    config_path = base / "config.json"

    if not config_path.exists():
        _crear_config_default(config_path)
        print(f"[INFO] config.json creado en {config_path}. Edítalo antes de continuar.")
        sys.exit(0)

    with open(config_path, encoding="utf-8") as f:
        return json.load(f)


def _crear_config_default(path: Path) -> None:
    config = {
        "_comentario": "Edita este archivo con tus credenciales. NO subas a GitHub.",
        "_version": "3.0",
        "portales": {
            "wherex": {
                "activo": True,
                "email": "",
                "password": ""
            },
            "portal_minero": {
                "activo": True,
                "usuario": "",
                "password": ""
            }
        },
        "rutas": {
            "output_dir": str(Path.home() / "Desktop" / "LicitaMonitor"),
            "log_dir":    str(Path.home() / "Desktop" / "LicitaMonitor" / "logs"),
        },
        "filtros": {
            "horas_atras": 24,
            "keywords": [
                "estructura", "caldereria", "caldería",
                "piping", "fabricacion", "fabricación",
                "montaje", "spool", "pipeline", "manifold",
                "tk", "estanques", "soldadura"
            ],
        },
        "selenium": {
            "headless": True,
            "timeout_segundos": 30,
            "max_reintentos": 3,
            "pausa_entre_paginas": 2,
        },
        "notificaciones": {
            "activar_email": False,
            "smtp_servidor": "smtp.gmail.com",
            "smtp_puerto": 587,
            "smtp_usuario": "",
            "smtp_password": "",
            "destinatarios": [],
        },
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)


# ═══════════════════════════════════════════════════════════════
# LOGGING
# ═══════════════════════════════════════════════════════════════

def configurar_logging(log_dir: str) -> logging.Logger:
    Path(log_dir).mkdir(parents=True, exist_ok=True)
    fecha    = datetime.now().strftime("%Y-%m-%d")
    log_file = Path(log_dir) / f"licitamonitor_{fecha}.log"

    logger = logging.getLogger("LicitaMonitor")
    if logger.handlers:
        return logger  # ya configurado

    logger.setLevel(logging.DEBUG)
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%Y-%m-%d %H:%M:%S")

    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)

    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)

    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger


# ═══════════════════════════════════════════════════════════════
# WEBDRIVER COMPARTIDO
# ═══════════════════════════════════════════════════════════════

def iniciar_driver(cfg_selenium: dict) -> webdriver.Chrome:
    options = webdriver.ChromeOptions()
    if cfg_selenium.get("headless", True):
        options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--lang=es-CL")
    # User-Agent real para evitar bloqueos por bot-detection (Portal Minero, etc.)
    options.add_argument(
        "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    # Logging de red para interceptar XHR (útil en SPAs como Wherex)
    options.set_capability("goog:loggingPrefs", {"performance": "ALL"})

    service = Service(ChromeDriverManager().install())
    driver  = webdriver.Chrome(service=service, options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver


def _guardar_screenshot(driver, nombre: str, logger) -> None:
    try:
        driver.save_screenshot(str(Path(nombre)))
        logger.debug("Screenshot: %s", nombre)
    except Exception:
        pass


def _es_relevante(texto: str, keywords: list) -> bool:
    """Verifica relevancia. Keywords de ≤2 chars usan word boundary para evitar falsos positivos."""
    txt = texto.lower()
    for kw in keywords:
        kw_l = kw.lower()
        if len(kw_l) <= 2:
            if re.search(r'\b' + re.escape(kw_l) + r'\b', txt):
                return True
        else:
            if kw_l in txt:
                return True
    return False


def _dentro_de_ventana(fecha_str: str, horas_atras: int, logger) -> bool:
    """True si la fecha está dentro de las últimas N horas. Retorna True si no parseable."""
    if not fecha_str or fecha_str.strip() in ("—", "-", ""):
        return True  # Sin fecha → incluir (portales sin columna de fecha)
    limite   = datetime.now() - timedelta(hours=horas_atras)
    formatos = [
        "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
        "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d",
        "%d-%m-%Y", "%d.%m.%Y",
    ]
    for fmt in formatos:
        try:
            return datetime.strptime(fecha_str.strip(), fmt) >= limite
        except ValueError:
            continue
    logger.debug("Fecha no parseable '%s' — incluida por defecto.", fecha_str)
    return True


# ═══════════════════════════════════════════════════════════════
# BASE PORTAL
# ═══════════════════════════════════════════════════════════════

class PortalBase(ABC):
    """Interfaz común para todos los scrapers de portal."""

    nombre_portal: str  = "Portal"
    color_badge:   str  = "1255A3"
    tiene_fechas:  bool = True   # False en portales sin columna de fecha

    def __init__(self, driver: webdriver.Chrome, cfg: dict, logger: logging.Logger):
        self.driver  = driver
        self.cfg     = cfg
        self.logger  = logger
        self.timeout = int(cfg.get("selenium", {}).get("timeout_segundos", 30))
        self.pausa   = float(cfg.get("selenium", {}).get("pausa_entre_paginas", 2))

    @abstractmethod
    def login(self) -> None: ...

    @abstractmethod
    def extraer_licitaciones(self, keywords: list, horas_atras: int) -> list: ...

    # Campos estándar que debe devolver cada licitación:
    # {
    #   "portal":           str,
    #   "id":               str,
    #   "comprador":        str,
    #   "descripcion":      str,
    #   "region":           str,
    #   "fecha_pub":        str,
    #   "fecha_cierre":     str,
    #   "link":             str,
    # }


# ═══════════════════════════════════════════════════════════════
# PORTAL: WHEREX
# ═══════════════════════════════════════════════════════════════

class PortalWherex(PortalBase):
    nombre_portal = "Wherex"
    color_badge   = C_BADGE_WHEREX

    LOGIN_URL = "https://login.wherex.com/?srv=system.wherex.com/secured/login_check&lang=es"
    BIDS_URL  = ("https://system.wherex.com/secured/supplier_user/"
                 "purchase-bids/all?bid_status=published&order=nearClosed")

    def login(self) -> None:
        self.logger.info("[Wherex] Iniciando sesión...")
        cfg_portal = self.cfg["portales"]["wherex"]
        email      = cfg_portal["email"]
        password   = cfg_portal["password"]

        self.driver.get(self.LOGIN_URL)
        wait = WebDriverWait(self.driver, self.timeout)

        # B5 fix: visibility_of_element_located espera que el campo sea interactuable
        # Wherex es SPA Vue.js — los campos se inyectan luego del JS inicial
        try:
            wait.until(EC.visibility_of_element_located((By.ID, "inputEmail")))
        except TimeoutException:
            # Fallback: Vue puede usar input[type='email'] o input[name='email']
            try:
                wait.until(EC.visibility_of_element_located(
                    (By.CSS_SELECTOR, "input[type='email']")))
            except TimeoutException:
                _guardar_screenshot(self.driver, "wherex_error_login.png", self.logger)
                raise RuntimeError("[Wherex] Página de login no cargó. Verifica conexión.")

        # Rellenar campos — probar ID primero, luego type
        for sel_email in [(By.ID, "inputEmail"), (By.CSS_SELECTOR, "input[type='email']")]:
            try:
                campo = self.driver.find_element(*sel_email)
                campo.clear(); campo.send_keys(email)
                break
            except NoSuchElementException:
                continue

        for sel_pass in [(By.ID, "inputPassword"), (By.CSS_SELECTOR, "input[type='password']")]:
            try:
                campo = self.driver.find_element(*sel_pass)
                campo.clear(); campo.send_keys(password)
                break
            except NoSuchElementException:
                continue

        # Botón login
        for sel_btn in [(By.ID, "loginEnter"), (By.CSS_SELECTOR, "button[type='submit']")]:
            try:
                btn = self.driver.find_element(*sel_btn)
                self.driver.execute_script("arguments[0].scrollIntoView(true);", btn)
                btn.click()
                break
            except NoSuchElementException:
                continue

        # B6 fix: redirect check más estricto
        try:
            wait.until(lambda d: (
                "/secured/supplier_user" in d.current_url.lower()
                and "login" not in d.current_url.lower()
            ))
            # B12 fix: espera extra para que Vue SPA termine de renderizar
            time.sleep(1.5)
            self.logger.info("[Wherex] Login exitoso → %s", self.driver.current_url)
        except TimeoutException:
            _guardar_screenshot(self.driver, "wherex_error_redirect.png", self.logger)
            raise RuntimeError("[Wherex] Login fallido: credenciales incorrectas o sin redirección.")

    def extraer_licitaciones(self, keywords: list, horas_atras: int) -> list:
        self.logger.info("[Wherex] Extrayendo licitaciones...")
        self.driver.get(self.BIDS_URL)

        wait         = WebDriverWait(self.driver, self.timeout)
        resultados   = []
        vistos: set  = set()
        pagina       = 1

        # B12 fix: espera extra para Vue.js renderizar la tabla
        time.sleep(2)
        try:
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr")))
        except TimeoutException:
            _guardar_screenshot(self.driver, "wherex_error_tabla.png", self.logger)
            self.logger.warning("[Wherex] Tabla de licitaciones no cargó.")
            return resultados

        while True:
            self.logger.info("[Wherex] Página %d...", pagina)
            try:
                wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr")))
            except TimeoutException:
                break

            filas = self.driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
            if not filas:
                break
            self.logger.debug("[Wherex] %d filas en página %d.", len(filas), pagina)

            stop = False
            for fila in filas:
                datos = self._parsear_fila(fila)
                if not datos:
                    continue

                if not _es_relevante(f"{datos['descripcion']} {datos['comprador']}", keywords):
                    continue

                # B7 fix: break inmediato al salir de rango para no saltar deduplicación
                if not _dentro_de_ventana(datos["fecha_pub"], horas_atras, self.logger):
                    stop = True
                    break

                clave = datos["id"] or datos["descripcion"][:50]
                if clave in vistos:
                    continue
                vistos.add(clave)
                resultados.append(datos)

            if stop:
                self.logger.info("[Wherex] Fechas fuera de rango — deteniendo paginación.")
                break

            # C1 fix: loop individual por selector en vez de coma-separados (inválido en Selenium)
            btn_next = None
            for sel in [
                "li.page-item:not(.disabled) a[aria-label='Next']",
                "li:not(.disabled) > a[rel='next']",
                ".pagination .next:not(.disabled) a",
                "button.btn-next:not([disabled])",
                "a[aria-label='Next']",
            ]:
                try:
                    btn_next = WebDriverWait(self.driver, 3).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, sel))
                    )
                    break
                except (TimeoutException, NoSuchElementException):
                    continue

            if not btn_next:
                self.logger.info("[Wherex] Sin más páginas.")
                break
            try:
                btn_next.click()
                time.sleep(self.pausa)
                pagina += 1
            except Exception as exc:
                self.logger.warning("[Wherex] Error en paginación: %s", exc)
                break

        self.logger.info("[Wherex] %d licitaciones encontradas.", len(resultados))
        return resultados

    def _parsear_fila(self, fila) -> dict | None:
        try:
            celdas = fila.find_elements(By.TAG_NAME, "td")
            if not fila.text.strip() or len(celdas) < 2:
                return None

            # Wherex tabla: col0=N°, col1=Título, col2=Comprador, col3=Producto/Servicio
            # col4=Monto, col5=Zona, col6=F.Publicación, col7=F.Cierre
            numero    = celdas[0].text.strip() if len(celdas) > 0 else "—"
            comprador = celdas[2].text.strip() if len(celdas) > 2 else "—"
            producto  = celdas[3].text.strip() if len(celdas) > 3 else \
                        celdas[1].text.strip() if len(celdas) > 1 else "—"
            zona      = celdas[5].text.strip() if len(celdas) > 5 else "—"
            fecha_pub = celdas[6].text.strip() if len(celdas) > 6 else "—"
            fecha_cie = celdas[7].text.strip() if len(celdas) > 7 else "—"

            # B8 fix: selectores de link más precisos
            link = ""
            for sel in ["a[href*='/purchase-bid']", "a[href*='/purchase-bids/']", "a[href]"]:
                try:
                    el = fila.find_element(By.CSS_SELECTOR, sel)
                    link = el.get_attribute("href") or ""
                    if link:
                        break
                except NoSuchElementException:
                    continue

            if not link and numero and numero != "—":
                link = (f"https://system.wherex.com/secured/supplier_user"
                        f"/purchase-bids/{numero}")

            return {
                "portal":       self.nombre_portal,
                "id":           numero,
                "comprador":    comprador,
                "descripcion":  producto,
                "region":       zona,
                "fecha_pub":    fecha_pub,
                "fecha_cierre": fecha_cie,
                "link":         link,
            }
        except StaleElementReferenceException:
            return None
        except Exception as exc:
            self.logger.debug("[Wherex] Error fila: %s", exc)
            return None


# ═══════════════════════════════════════════════════════════════
# PORTAL: PORTAL MINERO  (Atlassian Confluence)
# Login: https://www.portalminero.com/login.action
# Licitaciones: /display/serv/Listado+de+Proyectos
# ═══════════════════════════════════════════════════════════════

class PortalMinero(PortalBase):
    nombre_portal = "Portal Minero"
    color_badge   = C_BADGE_MINERO
    tiene_fechas  = False   # C4: la tabla no tiene columna de fecha

    # URL base sin parámetros de campaña — el portal redirige sólo al destino correcto
    LOGIN_URL  = "https://www.portalminero.com/login.action"
    BIDS_URL   = "https://www.portalminero.com/display/serv/Listado+de+Proyectos"
    BASE_URL   = "https://www.portalminero.com"

    def login(self) -> None:
        self.logger.info("[Portal Minero] Iniciando sesión (Confluence)...")
        cfg_portal = self.cfg["portales"]["portal_minero"]
        usuario    = cfg_portal["usuario"]
        password   = cfg_portal["password"]

        self.driver.get(self.LOGIN_URL)
        wait = WebDriverWait(self.driver, self.timeout)

        # ── Esperar formulario Confluence (C1: sin comas en selector) ────
        found = False
        for sel in ["#os_username", "input[name='os_username']", "#username"]:
            try:
                wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, sel)))
                found = True
                break
            except TimeoutException:
                continue
        if not found:
            _guardar_screenshot(self.driver, "pm_error_login_page.png", self.logger)
            raise RuntimeError("[Portal Minero] Página de login no cargó en %ds." % self.timeout)

        # ── Campos (Confluence estándar + fallbacks) ─────────────────────
        CAMPOS_USER = [
            (By.ID,   "os_username"),
            (By.NAME, "os_username"),
            (By.ID,   "username"),
            (By.NAME, "username"),
            (By.CSS_SELECTOR, "input[type='text']:not([disabled])"),
        ]
        CAMPOS_PASS = [
            (By.ID,   "os_password"),
            (By.NAME, "os_password"),
            (By.ID,   "password"),
            (By.NAME, "password"),
            (By.CSS_SELECTOR, "input[type='password']:not([disabled])"),
        ]
        CAMPOS_BTN = [
            (By.ID,   "loginButton"),
            (By.CSS_SELECTOR, "input#loginButton"),
            (By.CSS_SELECTOR, "button#loginButton"),
            (By.CSS_SELECTOR, "input[type='submit']"),
            (By.CSS_SELECTOR, "button[type='submit']"),
            (By.XPATH, "//input[@value='Iniciar sesión']"),
            (By.XPATH, "//input[@value='Log In']"),
            (By.XPATH, "//button[normalize-space()='Iniciar sesión']"),
        ]

        campo_user = self._encontrar_elemento(CAMPOS_USER)
        campo_pass = self._encontrar_elemento(CAMPOS_PASS)

        if not campo_user or not campo_pass:
            _guardar_screenshot(self.driver, "pm_error_campos.png", self.logger)
            raise RuntimeError("[Portal Minero] No se encontraron los campos usuario/contraseña.")

        # Limpiar y rellenar con JS para evitar autocompletado problemático
        self.driver.execute_script("arguments[0].value = '';", campo_user)
        campo_user.send_keys(usuario)
        self.driver.execute_script("arguments[0].value = '';", campo_pass)
        campo_pass.send_keys(password)

        btn = self._encontrar_elemento(CAMPOS_BTN)
        if btn:
            self.driver.execute_script("arguments[0].click();", btn)
        else:
            campo_pass.submit()

        # ── Detectar login exitoso ────────────────────────────────────────
        # Confluence redirige fuera de /login.action al autenticar
        try:
            wait.until(lambda d: (
                "login.action" not in d.current_url
                and "login" not in d.current_url.lower().split("?")[0]
            ))
            self.logger.info("[Portal Minero] Login exitoso → %s", self.driver.current_url)
        except TimeoutException:
            # Verificar si hay mensaje de error visible
            try:
                err = self.driver.find_element(
                    By.CSS_SELECTOR,
                    ".aui-message-error, #loginErrorDiv, .error, [class*='error-message']"
                )
                msg = err.text.strip()
            except Exception:
                msg = "sin mensaje"
            _guardar_screenshot(self.driver, "pm_error_redirect.png", self.logger)
            raise RuntimeError(
                f"[Portal Minero] Login fallido — {msg}. "
                "Verifica usuario y contraseña en config.json."
            )

    def extraer_licitaciones(self, keywords: list, horas_atras: int) -> list:
        self.logger.info("[Portal Minero] Accediendo a listado de proyectos...")
        self.driver.get(self.BIDS_URL)

        wait        = WebDriverWait(self.driver, self.timeout)
        resultados  = []
        vistos: set = set()
        pagina      = 1

        # ── Esperar contenido Confluence ─────────────────────────────────
        # El listado de proyectos en Confluence suele estar en #main-content
        # o dentro de tablas/macros de Confluence
        try:
            wait.until(EC.presence_of_element_located(
                (By.CSS_SELECTOR,
                 "#main-content, #content, .wiki-content, table.confluenceTable, "
                 ".page-metadata-modification-info")
            ))
            time.sleep(2)  # Dar tiempo al JS de Confluence
        except TimeoutException:
            _guardar_screenshot(self.driver, "pm_error_lista.png", self.logger)
            self.logger.warning("[Portal Minero] Página de proyectos no cargó. ¿Sesión expirada?")
            return resultados

        while True:
            self.logger.info("[Portal Minero] Página %d...", pagina)

            items = self._obtener_items_pagina()
            if not items:
                self.logger.info("[Portal Minero] Sin ítems detectados en página %d.", pagina)
                break

            self.logger.debug("[Portal Minero] %d filas detectadas.", len(items))

            for item in items:
                datos = self._parsear_item(item)
                if not datos:
                    continue
                if not _es_relevante(
                    f"{datos['descripcion']} {datos['comprador']} {datos['region']}",
                    keywords
                ):
                    continue
                # C4: Portal Minero no tiene columna de fecha — saltar filtro temporal
                if self.tiene_fechas and not _dentro_de_ventana(
                    datos["fecha_pub"], horas_atras, self.logger
                ):
                    continue

                clave = datos["id"] or datos["descripcion"][:50]
                if clave in vistos:
                    continue
                vistos.add(clave)
                resultados.append(datos)

            # Paginación Confluence (macro de tabla paginada o "siguiente" link)
            siguiente = self._siguiente_pagina(wait)
            if not siguiente:
                self.logger.info("[Portal Minero] Sin más páginas.")
                break
            try:
                self.driver.execute_script("arguments[0].click();", siguiente)
                time.sleep(self.pausa)
                pagina += 1
            except Exception as exc:
                self.logger.warning("[Portal Minero] Error avanzando página: %s", exc)
                break

        self.logger.info("[Portal Minero] %d licitaciones relevantes.", len(resultados))
        return resultados

    def _obtener_items_pagina(self) -> list:
        """
        Portal Minero usa Confluence — el contenido está en tablas .confluenceTable
        o dentro de #main-content. Intenta múltiples selectores en orden de prioridad.
        """
        SELECTORES = [
            # Tablas Confluence (más específico primero)
            "table.confluenceTable tbody tr",
            "#main-content table tbody tr",
            ".wiki-content table tbody tr",
            "#content table tbody tr",
            # Fallback genérico
            "table tbody tr",
            # Posibles cards/divs si usan macro visual
            ".project-item",
            ".licitacion-row",
            "ul.project-list li",
        ]
        for sel in SELECTORES:
            items = self.driver.find_elements(By.CSS_SELECTOR, sel)
            # B9: no filtrar por <th> — solo quitar filas completamente vacías
            items = [i for i in items if i.text.strip()]
            if items:
                self.logger.debug("[Portal Minero] Selector activo: %s (%d items)", sel, len(items))
                return items
        return []

    def _parsear_item(self, item) -> dict | None:
        """
        C2: Estructura real de la tabla de Portal Minero (4 columnas):
          col 0 — Nombre del proyecto (con <a> al detalle)
          col 1 — País
          col 2 — Región
          col 3 — Ver Detalles (botón/link al detalle)
        Sin columnas de fecha — tiene_fechas = False.
        """
        try:
            texto = item.text.strip()
            if not texto or len(texto) < 4:
                return None

            celdas = item.find_elements(By.TAG_NAME, "td")
            n = len(celdas)

            if n == 0:
                return None  # fila de encabezado con <th>

            # Mapeo correcto según estructura real de 4 columnas
            desc      = celdas[0].text.strip() if n > 0 else texto[:120]
            pais      = celdas[1].text.strip() if n > 1 else "—"
            region    = celdas[2].text.strip() if n > 2 else "—"
            # col 3 es solo "Ver Detalles" — no tiene info adicional útil

            # ── Link: puede estar en col 0 (nombre) o col 3 (Ver Detalles) ──
            link = ""
            for sel in ["a[href*='/display/']", "a[href*='portalminero']", "a[href]"]:
                try:
                    el   = item.find_element(By.CSS_SELECTOR, sel)
                    href = el.get_attribute("href") or ""
                    if href and not href.startswith("javascript"):
                        link = href if href.startswith("http") else self.BASE_URL + href
                        break
                except NoSuchElementException:
                    continue

            return {
                "portal":       self.nombre_portal,
                "id":           "—",
                "comprador":    pais,        # usamos "comprador" para País (más visible en Excel)
                "descripcion":  desc,
                "region":       region,
                "fecha_pub":    "—",
                "fecha_cierre": "—",
                "link":         link,
            }

        except StaleElementReferenceException:
            return None
        except Exception as exc:
            self.logger.debug("[Portal Minero] Error parseando item: %s", exc)
            return None

    def _extraer_attr(self, parent, selector: str, attr: str) -> str:
        try:
            el = parent.find_element(By.CSS_SELECTOR, selector)
            return el.text.strip() if attr == "text" else (el.get_attribute(attr) or "")
        except Exception:
            return ""

    def _siguiente_pagina(self, wait) -> object | None:
        """
        C3: Portal Minero usa paginación JS vía javascript:pagina(N,1).
        Detectamos el link de la página siguiente y lo ejecutamos directamente.
        Devuelve el elemento clickable o None si no hay más páginas.
        """
        # 1. Buscar links con onclick o href tipo javascript:pagina(N,1)
        try:
            enlaces = self.driver.find_elements(By.CSS_SELECTOR, "a")
            pagina_actual = None
            proximo_pagina = None

            for a in enlaces:
                href = a.get_attribute("href") or ""
                onclick = a.get_attribute("onclick") or ""
                texto_a = a.text.strip()

                # Detectar patrón javascript:pagina(N,1)
                for fuente in [href, onclick]:
                    m = re.search(r'pagina\((\d+)\s*,\s*1\)', fuente)
                    if m:
                        # El link con clase "active" o "current" es la página actual
                        clases = (a.get_attribute("class") or "").lower()
                        if "active" in clases or "current" in clases or "selected" in clases:
                            pagina_actual = int(m.group(1))
                        else:
                            # Guardar el candidato a "siguiente" (mayor número)
                            n = int(m.group(1))
                            if proximo_pagina is None or n > proximo_pagina[0]:
                                proximo_pagina = (n, a, fuente)

            if proximo_pagina:
                n, el, fuente = proximo_pagina
                # Si tenemos página actual, solo avanzar exactamente 1
                if pagina_actual is not None and n != pagina_actual + 1:
                    # Buscar específicamente la página actual + 1
                    for a in enlaces:
                        for fuente2 in [a.get_attribute("href") or "", a.get_attribute("onclick") or ""]:
                            m = re.search(r'pagina\((\d+)\s*,\s*1\)', fuente2)
                            if m and int(m.group(1)) == pagina_actual + 1:
                                return a
                    return None
                return el
        except Exception as exc:
            self.logger.debug("[Portal Minero] _siguiente_pagina JS: %s", exc)

        # 2. Fallback: selectores estándar de paginación
        for selector in [
            "a[aria-label='Next']:not(.disabled)",
            "a[rel='next']",
            ".pagination li:not(.disabled) a.next",
            "a.siguiente:not(.disabled)",
            "a[title*='siguiente' i]",
            "a[title*='next' i]",
            "li.next:not(.disabled) a",
        ]:
            try:
                el = self.driver.find_element(By.CSS_SELECTOR, selector)
                if el.is_displayed() and el.is_enabled():
                    return el
            except NoSuchElementException:
                continue
        return None

    def _encontrar_elemento(self, selectors: list):
        for by, sel in selectors:
            try:
                el = self.driver.find_element(by, sel)
                if el.is_displayed():
                    return el
            except Exception:
                continue
        return None


# ═══════════════════════════════════════════════════════════════
# GENERACIÓN DE EXCEL — PALETA OSCURA
# ═══════════════════════════════════════════════════════════════

def _borde_oscuro():
    lado = Side(style="thin", color=C_BORDER)
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)

def _font(bold=False, size=10, color=C_TEXT_DATA, italic=False, underline=None):
    return Font(bold=bold, size=size, color=color, italic=italic,
                underline=underline or "none")


def generar_excel(
    licitaciones: list,
    output_dir: str,
    keywords: list,
    horas_atras: int,
    logger: logging.Logger,
) -> str:
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    ts       = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filepath = Path(output_dir) / f"LicitaMonitor_Soldesp_{ts}.xlsx"

    wb = openpyxl.Workbook()

    _hoja_licitaciones(wb.active, licitaciones, keywords, horas_atras)

    # Hojas por portal
    portales_presentes = sorted({l["portal"] for l in licitaciones})
    for portal in portales_presentes:
        subset = [l for l in licitaciones if l["portal"] == portal]
        hoja_nombre = portal[:31]
        ws_p = wb.create_sheet(hoja_nombre)
        _hoja_licitaciones(ws_p, subset, keywords, horas_atras, subtitulo=f"Portal: {portal}")

    _hoja_resumen(wb.create_sheet("Resumen"), licitaciones, keywords, horas_atras, filepath.name)

    try:
        wb.save(str(filepath))
        logger.info("Excel guardado: %s", filepath)
    except PermissionError:
        alt = filepath.with_stem(filepath.stem + "_v2")
        wb.save(str(alt))
        logger.warning("Archivo en uso — guardado como: %s", alt)
        return str(alt)

    return str(filepath)


def _hoja_licitaciones(ws, licitaciones: list, keywords: list, horas_atras: int,
                       subtitulo: str | None = None) -> None:
    ws.sheet_view.showGridLines = False

    borde   = _borde_oscuro()
    aln_c   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    aln_i   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── Fila 1: Título principal ────────────────────────────────────────
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value     = (f"LICITAMONITOR — SOLDESP  ·  "
                   f"{datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  "
                   f"Período: últimas {horas_atras}h")
    c.font      = _font(bold=True, size=13, color=C_TEXT_HEADER)
    c.fill      = _fill(C_HEADER_BG)
    c.alignment = aln_c
    ws.row_dimensions[1].height = 34

    # ── Fila 2: Subtítulo / filtros ─────────────────────────────────────
    ws.merge_cells("A2:I2")
    c2 = ws["A2"]
    if subtitulo:
        c2.value = subtitulo
    else:
        c2.value = "Filtros: " + " · ".join(kw.capitalize() for kw in keywords)
    c2.font      = _font(italic=True, size=9, color=C_TEXT_MUTED)
    c2.fill      = _fill(C_SUBHEADER_BG)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Fila 3: Cabeceras ────────────────────────────────────────────────
    headers = ["Portal", "ID / N°", "Comprador / Mandante",
               "Descripción / Servicio", "Región / Zona",
               "Fecha Publicación", "Fecha Cierre", "Link Directo", "Estado"]
    anchos  = [18,       12,          28,
               50,                      16,
               20,                 20,              24,              12]

    for col_i, (hdr, ancho) in enumerate(zip(headers, anchos), start=1):
        c = ws.cell(row=3, column=col_i, value=hdr)
        c.font      = _font(bold=True, size=10, color=C_TEXT_HEADER)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = aln_c
        c.border    = borde
        ws.column_dimensions[get_column_letter(col_i)].width = ancho
    ws.row_dimensions[3].height = 26

    # ── Filas de datos ───────────────────────────────────────────────────
    if not licitaciones:
        ws.merge_cells("A4:I4")
        c = ws["A4"]
        c.value     = f"No se encontraron licitaciones en las últimas {horas_atras} horas."
        c.font      = _font(italic=True, color=C_TEXT_MUTED)
        c.fill      = _fill(C_ROW_ODD)
        c.alignment = aln_c
        ws.row_dimensions[4].height = 28
    else:
        portales_colores = {}
        for row_i, lic in enumerate(licitaciones, start=4):
            bg = C_ROW_EVEN if row_i % 2 == 0 else C_ROW_ODD
            fill_data = _fill(bg)

            # Celda portal con color de badge
            portal_color = _color_portal(lic["portal"])
            c_portal = ws.cell(row=row_i, column=1, value=lic["portal"])
            c_portal.font      = _font(bold=True, size=9, color=C_TEXT_HEADER)
            c_portal.fill      = _fill(portal_color)
            c_portal.alignment = aln_c
            c_portal.border    = borde

            # Resto de columnas
            valores = [
                lic["id"],
                lic["comprador"],
                lic["descripcion"],
                lic["region"],
                lic["fecha_pub"],
                lic["fecha_cierre"],
                None,            # placeholder Link
                "Vigente",
            ]

            for col_i, valor in enumerate(valores, start=2):
                c = ws.cell(row=row_i, column=col_i, value=valor)
                c.font      = _font(size=9, color=C_TEXT_DATA)
                c.fill      = fill_data
                c.border    = borde
                c.alignment = aln_c if col_i in (2, 5, 6, 7, 9) else aln_i

            # Hipervínculo en columna 8
            lc = ws.cell(row=row_i, column=8)
            if lic["link"]:
                lc.value     = "Ver licitación →"
                lc.hyperlink = lic["link"]
                lc.font      = _font(size=9, color=C_LINK, underline="single")
            else:
                lc.value = "Sin enlace"
                lc.font  = _font(size=9, color=C_TEXT_MUTED)
            lc.fill      = fill_data
            lc.border    = borde
            lc.alignment = aln_c

            ws.row_dimensions[row_i].height = 26

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:I{3 + len(licitaciones)}"


def _hoja_resumen(ws, licitaciones: list, keywords: list,
                  horas_atras: int, nombre_archivo: str) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    borde = _borde_oscuro()

    # Título
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value     = "RESUMEN DE EJECUCIÓN — LicitaMonitor Soldesp"
    c.font      = _font(bold=True, size=12, color=C_TEXT_HEADER)
    c.fill      = _fill(C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Estadísticas globales
    portales = sorted({l["portal"] for l in licitaciones})
    stats = [
        ("Fecha de ejecución",            datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("Período analizado",             f"Últimas {horas_atras} horas"),
        ("─── Resultados globales ───",   ""),
        ("Total licitaciones encontradas", len(licitaciones)),
        ("Compradores distintos",          len({l["comprador"] for l in licitaciones})),
        ("Regiones distintas",             len({l["region"]    for l in licitaciones if l["region"] != "—"})),
        ("─── Por portal ───",            ""),
    ]
    for p in portales:
        cnt = sum(1 for l in licitaciones if l["portal"] == p)
        stats.append((f"  · {p}", cnt))

    stats += [
        ("─── Filtros aplicados ───",     ""),
        ("Keywords",                      " · ".join(kw.capitalize() for kw in keywords)),
        ("Archivo generado",              nombre_archivo),
    ]

    for row_i, (label, valor) in enumerate(stats, start=2):
        c_label = ws.cell(row=row_i, column=1, value=label)
        c_valor = ws.cell(row=row_i, column=2, value=valor)

        es_seccion = label.startswith("───")
        bg = C_SUBHEADER_BG if es_seccion else (C_ROW_EVEN if row_i % 2 == 0 else C_ROW_ODD)

        c_label.font = _font(bold=True, size=10,
                              color=(C_TEXT_MUTED if es_seccion else C_TEXT_DATA))
        c_valor.font = _font(size=10, color=C_TEXT_DATA)

        for c in (c_label, c_valor):
            c.fill      = _fill(bg)
            c.border    = borde
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row_i].height = 22


def _color_portal(nombre: str) -> str:
    tabla = {
        "Wherex":        C_BADGE_WHEREX,
        "Portal Minero": C_BADGE_MINERO,
    }
    return tabla.get(nombre, "1B3A6B")


# ═══════════════════════════════════════════════════════════════
# NOTIFICACIÓN EMAIL
# ═══════════════════════════════════════════════════════════════

def enviar_email(cfg_notif: dict, filepath: str, n_lics: int, logger: logging.Logger) -> None:
    if not cfg_notif.get("activar_email", False) or not cfg_notif.get("destinatarios"):
        return
    try:
        msg            = MIMEMultipart()
        msg["From"]    = cfg_notif["smtp_usuario"]
        msg["To"]      = ", ".join(cfg_notif["destinatarios"])
        msg["Subject"] = (f"LicitaMonitor Soldesp — {n_lics} licitaciones "
                          f"({datetime.now().strftime('%d/%m/%Y')})")

        cuerpo = (
            f"Se encontraron <b>{n_lics}</b> licitaciones relevantes.<br>"
            f"Archivo adjunto: <code>{Path(filepath).name}</code><br><br>"
            f"<i>Generado automáticamente por LicitaMonitor Soldesp</i>"
        )
        msg.attach(MIMEText(cuerpo, "html", "utf-8"))

        with open(filepath, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition",
                        f'attachment; filename="{Path(filepath).name}"')
        msg.attach(part)

        with smtplib.SMTP(cfg_notif["smtp_servidor"], int(cfg_notif["smtp_puerto"])) as s:
            s.ehlo(); s.starttls()
            s.login(cfg_notif["smtp_usuario"], cfg_notif["smtp_password"])
            s.sendmail(cfg_notif["smtp_usuario"],
                       cfg_notif["destinatarios"], msg.as_string())

        logger.info("Email enviado a: %s", ", ".join(cfg_notif["destinatarios"]))
    except Exception as exc:
        logger.warning("Email no enviado: %s", exc)


# ═══════════════════════════════════════════════════════════════
# ORQUESTADOR PRINCIPAL
# ═══════════════════════════════════════════════════════════════

PORTAL_MAP = {
    "wherex":        PortalWherex,
    "portal_minero": PortalMinero,
}

def ejecutar_proceso(
    logger: logging.Logger | None = None,
    portales_activos: list | None = None,
) -> tuple[str | None, int]:
    """
    Ejecuta scraping en todos los portales activos y genera un Excel unificado.
    Retorna (ruta_excel, total_licitaciones).
    """
    cfg     = cargar_config()
    log_dir = cfg["rutas"].get("log_dir", str(Path.home() / "Desktop" / "LicitaMonitor" / "logs"))
    if logger is None:
        logger = configurar_logging(log_dir)

    keywords    = cfg["filtros"]["keywords"]
    horas_atras = int(cfg["filtros"].get("horas_atras", 24))
    output_dir  = cfg["rutas"]["output_dir"]
    cfg_sel     = cfg.get("selenium", {})
    reintentos  = int(cfg_sel.get("max_reintentos", 3))

    # Determinar portales a ejecutar
    portales_cfg = cfg.get("portales", {})
    if portales_activos is None:
        portales_activos = [k for k, v in portales_cfg.items() if v.get("activo", True)]

    logger.info("=" * 60)
    logger.info("  LICITAMONITOR — %s", datetime.now().strftime("%d/%m/%Y %H:%M"))
    logger.info("  Portales activos: %s", ", ".join(portales_activos))
    logger.info("  Keywords: %s", ", ".join(keywords))
    logger.info("=" * 60)

    todas_licitaciones: list = []
    driver = None

    for nombre_portal in portales_activos:
        cls = PORTAL_MAP.get(nombre_portal)
        if not cls:
            logger.warning("Portal desconocido: %s — omitido.", nombre_portal)
            continue

        for intento in range(1, reintentos + 1):
            try:
                if intento > 1:
                    logger.info("Reintento %d/%d para %s...", intento, reintentos, nombre_portal)
                    time.sleep(5 * intento)

                driver  = iniciar_driver(cfg_sel)
                portal  = cls(driver, cfg, logger)
                portal.login()
                lics    = portal.extraer_licitaciones(keywords, horas_atras)
                todas_licitaciones.extend(lics)
                logger.info("[%s] %d licitaciones agregadas.", nombre_portal, len(lics))
                break

            except RuntimeError as exc:
                logger.error("Error en %s: %s", nombre_portal, exc)
            except WebDriverException as exc:
                logger.error("WebDriver error en %s: %s", nombre_portal, exc)
            except Exception as exc:
                logger.exception("Error inesperado en %s: %s", nombre_portal, exc)
            finally:
                if driver:
                    try:
                        driver.quit()
                    except Exception:
                        pass
                    driver = None

    archivo = None
    if todas_licitaciones or True:  # Generar Excel siempre (aunque esté vacío)
        archivo = generar_excel(todas_licitaciones, output_dir, keywords, horas_atras, logger)

    if archivo:
        logger.info("")
        logger.info("╔══════════════════════════════════════════════════════════╗")
        logger.info("║  PROCESO COMPLETADO                                      ║")
        logger.info("╠══════════════════════════════════════════════════════════╣")
        logger.info("║  Total licitaciones: %-36d║", len(todas_licitaciones))
        logger.info("║  Archivo: %-48s║", Path(archivo).name[:48])
        logger.info("╚══════════════════════════════════════════════════════════╝")
        enviar_email(cfg.get("notificaciones", {}), archivo, len(todas_licitaciones), logger)

    return archivo, len(todas_licitaciones)


def main() -> None:
    cfg     = cargar_config()
    log_dir = cfg["rutas"].get("log_dir", str(Path.home() / "Desktop" / "LicitaMonitor" / "logs"))
    logger  = configurar_logging(log_dir)
    ejecutar_proceso(logger)


if __name__ == "__main__":
    main()
